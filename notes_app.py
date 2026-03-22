import streamlit as st
import pandas as pd
import math
import io
from datetime import datetime

# Vérification des modules optionnels
try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from reportlab.lib.pagesizes import A4
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

st.set_page_config(
    page_title="Suivi des Notes",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700;800;900&family=JetBrains+Mono:wght@400;600&display=swap');

:root {
    --blue:    #1565C0;
    --blue2:   #0D47A1;
    --sky:     #E3F2FD;
    --gold:    #F9A825;
    --green:   #2E7D32;
    --red:     #C62828;
    --orange:  #E65100;
    --bg:      #F0F4F8;
    --card:    #FFFFFF;
    --text:    #0D1B2A;
    --muted:   #546E7A;
    --border:  #CFD8DC;
}

* { font-family: 'Outfit', sans-serif; color: var(--text); }

.stApp {
    background: var(--bg);
    background-image: radial-gradient(circle at 15% 15%, rgba(21,101,192,0.06) 0%, transparent 60%),
                      radial-gradient(circle at 85% 85%, rgba(249,168,37,0.05) 0%, transparent 60%);
}

/* ── Header ── */
.app-header {
    background: linear-gradient(135deg, var(--blue2) 0%, var(--blue) 60%, #1976D2 100%);
    border-radius: 20px;
    padding: 30px 36px 26px;
    margin-bottom: 24px;
    position: relative;
    overflow: hidden;
    box-shadow: 0 8px 32px rgba(13,71,161,0.28);
}
.app-header::before {
    content: '🎓';
    position: absolute; right: 32px; top: 50%;
    transform: translateY(-50%);
    font-size: 90px; opacity: 0.10; line-height: 1;
}
.app-header::after {
    content: '';
    position: absolute; bottom: 0; left: 0; right: 0; height: 4px;
    background: linear-gradient(90deg, var(--gold), transparent);
}
.app-tag   { font-size:12px; font-weight:700; letter-spacing:4px; color:rgba(255,255,255,0.55); text-transform:uppercase; margin-bottom:6px; }
.app-title { font-size:32px; font-weight:900; color:#fff; line-height:1.1; margin:0; }
.app-sub   { font-size:14px; color:rgba(255,255,255,0.6); margin-top:6px; }

/* ── Cards ── */
.card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 16px;
    padding: 24px;
    margin-bottom: 18px;
    box-shadow: 0 2px 12px rgba(13,71,161,0.06);
}
.card-title {
    font-size: 16px; font-weight: 800; color: var(--blue2);
    border-bottom: 2px solid var(--sky);
    padding-bottom: 10px; margin-bottom: 18px;
    letter-spacing: 0.3px;
}

/* ── Stat boxes ── */
.stat-grid { display:flex; gap:12px; flex-wrap:wrap; margin-bottom:18px; }
.stat-box {
    background: var(--sky); border-radius:12px; padding:14px 20px;
    border-left: 4px solid var(--blue); flex:1; min-width:120px;
}
.stat-box.gold  { background:#FFF8E1; border-color:var(--gold); }
.stat-box.green { background:#E8F5E9; border-color:var(--green); }
.stat-box.red   { background:#FFEBEE; border-color:var(--red); }
.stat-label { font-size:11px; font-weight:700; letter-spacing:2px; text-transform:uppercase; color:var(--muted); margin-bottom:4px; }
.stat-value { font-size:26px; font-weight:900; color:var(--blue2); line-height:1; }
.stat-box.gold  .stat-value { color:#7B5800; }
.stat-box.green .stat-value { color:var(--green); }
.stat-box.red   .stat-value { color:var(--red); }

/* ── Mention badge ── */
.mention {
    display:inline-block; padding:3px 12px; border-radius:20px;
    font-size:12px; font-weight:700; letter-spacing:0.5px;
}
.mention-excellent { background:#E8F5E9; color:#1B5E20; border:1px solid #A5D6A7; }
.mention-bien      { background:#E3F2FD; color:#0D47A1; border:1px solid #90CAF9; }
.mention-ab        { background:#FFF8E1; color:#7B5800; border:1px solid #FFE082; }
.mention-passable  { background:#FFF3E0; color:#E65100; border:1px solid #FFCC80; }
.mention-insuffisant{ background:#FFEBEE; color:#B71C1C; border:1px solid #EF9A9A; }

/* ── Progress bar ── */
.prog-bar-wrap { background:#E0E7EF; border-radius:999px; height:8px; margin:4px 0; overflow:hidden; }
.prog-bar { height:8px; border-radius:999px; }
.prog-up   { background:linear-gradient(90deg,#43A047,#66BB6A); }
.prog-down { background:linear-gradient(90deg,#EF5350,#E57373); }
.prog-flat { background:linear-gradient(90deg,#FFA726,#FFB74D); }

/* ── Table ── */
.stDataFrame { border-radius:12px !important; overflow:hidden; }

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    background: var(--sky); border-radius:12px; padding:4px; gap:4px; border:none; margin-bottom:18px;
}
.stTabs [data-baseweb="tab"] {
    border-radius:8px; font-weight:700; font-size:13px; color:var(--muted);
    padding:9px 18px; border:none; background:transparent;
}
.stTabs [aria-selected="true"] { background:var(--blue) !important; color:white !important; }
.stTabs [data-baseweb="tab-border"] { display:none; }

/* ── Buttons ── */
.stButton > button {
    background:linear-gradient(135deg,var(--blue),var(--blue2)) !important;
    color:white !important; border:none !important; border-radius:10px !important;
    font-weight:700 !important; font-size:14px !important; padding:11px 28px !important;
    box-shadow:0 4px 14px rgba(13,71,161,0.25) !important; width:100%;
}
.stDownloadButton > button {
    background:white !important; border-radius:10px !important;
    font-weight:700 !important; font-size:13px !important;
    padding:9px 20px !important; width:100%; margin-top:6px;
}
.dl-excel .stDownloadButton > button { color:var(--green) !important; border:2px solid var(--green) !important; }
.dl-pdf   .stDownloadButton > button { color:var(--red) !important;   border:2px solid var(--red)   !important; }

/* ── Inputs ── */
.stSelectbox label, .stFileUploader label { font-weight:700 !important; color:var(--text) !important; }

/* ── Metrics ── */
[data-testid="metric-container"] {
    background:var(--sky); border-radius:10px; padding:12px 16px !important; border:1px solid var(--border);
}
[data-testid="stMetricLabel"] p  { color:#333 !important; font-weight:700 !important; font-size:12px !important; }
[data-testid="stMetricValue"] div{ color:var(--blue2) !important; font-weight:900 !important; font-size:22px !important; }

/* ── Upload zone ── */
[data-testid="stFileUploader"] {
    background:var(--sky) !important; border:2px dashed var(--blue) !important;
    border-radius:12px !important; padding:12px !important;
}
/* Browse files button → blanc */
[data-testid="stFileUploader"] button {
    background: var(--blue) !important;
    color: white !important;
    border: none !important;
    font-weight: 700 !important;
}
[data-testid="stFileUploader"] button:hover {
    background: var(--blue2) !important;
}
/* Nom du fichier uploadé → noir */
[data-testid="stFileUploader"] [data-testid="stMarkdownContainer"] p,
[data-testid="stFileUploader"] span,
[data-testid="stFileUploader"] small,
[data-testid="uploadedFileData"] span {
    color: #1A1A1A !important;
    font-weight: 600 !important;
}

/* ── Footer ── */
.app-footer { text-align:center; padding:18px; color:var(--muted); font-size:12px; margin-top:24px; border-top:1px solid var(--border); }
.badge { display:inline-block; background:var(--sky); color:var(--blue2); font-size:11px; font-weight:700; padding:3px 10px; border-radius:20px; border:1px solid var(--border); }

#MainMenu, footer, header { visibility:hidden; }
.block-container { padding-top: 2rem; }
</style>
""", unsafe_allow_html=True)

# ── Helpers ───────────────────────────────────────────────────────────────────
def get_mention(moyenne):
    if moyenne >= 16:   return "Excellent",   "mention-excellent"
    elif moyenne >= 14: return "Bien",         "mention-bien"
    elif moyenne >= 12: return "Assez Bien",   "mention-ab"
    elif moyenne >= 10: return "Passable",     "mention-passable"
    else:               return "Insuffisant",  "mention-insuffisant"

def get_trend(d1, d2, d3=None):
    """Retourne tendance globale. Si d3=None, compare D2 vs D1."""
    last = d3 if d3 is not None and pd.notna(d3) else d2
    diff = last - d1
    if diff > 1:    return "En progres",  "green"
    elif diff < -1: return "En baisse",   "red"
    else:           return "Stable",      "orange"

def progress_bar(val, max_val=20, trend="flat"):
    pct = min(100, (val / max_val) * 100)
    cls = "prog-up" if trend=="green" else ("prog-down" if trend=="red" else "prog-flat")
    return f'<div class="prog-bar-wrap"><div class="prog-bar {cls}" style="width:{pct}%"></div></div>'

def make_excel_bulletin(df_result):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulletin de Notes"

    blue_fill  = PatternFill("solid", fgColor="1565C0")
    blue2_fill = PatternFill("solid", fgColor="0D47A1")
    sky_fill   = PatternFill("solid", fgColor="E3F2FD")
    gold_fill  = PatternFill("solid", fgColor="FFF8E1")
    green_fill = PatternFill("solid", fgColor="E8F5E9")
    red_fill   = PatternFill("solid", fgColor="FFEBEE")
    thin = Side(style="thin", color="CFD8DC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def wc(r, c, v, fill=None, font=None, align="center"):
        x = ws.cell(row=r, column=c, value=v)
        if fill: x.fill = fill
        if font: x.font = font
        x.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        x.border = bdr
        return x

    # Titre
    ws.merge_cells("A1:H1")
    wc(1,1, "🎓 BULLETIN DE NOTES — SUIVI DE PROGRESSION",
       fill=blue_fill, font=Font(bold=True, color="FFFFFF", size=14))
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    wc(2,1, f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
       font=Font(italic=True, color="546E7A", size=10))
    ws.row_dimensions[2].height = 18

    # En-têtes
    headers = ["N°", "Nom Étudiant", "Devoir 1", "Devoir 2", "Devoir 3", "Moyenne", "Mention", "Tendance"]
    for c, h in enumerate(headers, 1):
        wc(3, c, h, fill=blue2_fill, font=Font(bold=True, color="FFFFFF", size=11))
    ws.row_dimensions[3].height = 22

    # Données
    for i, row in enumerate(df_result.itertuples(), 1):
        r = i + 3
        mention, _ = get_mention(row.Moyenne)
        trend_txt, trend_color = get_trend(row.Devoir1, row.Devoir2, row.Devoir3)

        # Couleur ligne selon mention
        if row.Moyenne >= 14:   row_fill = green_fill
        elif row.Moyenne >= 10: row_fill = sky_fill
        else:                   row_fill = red_fill

        wc(r, 1, i,             fill=row_fill, font=Font(bold=True))
        wc(r, 2, row.Etudiant,  fill=row_fill, font=Font(bold=True), align="left")
        wc(r, 3, row.Devoir1,   fill=row_fill)
        wc(r, 4, row.Devoir2,   fill=row_fill)
        wc(r, 5, row.Devoir3,   fill=row_fill)
        wc(r, 6, round(row.Moyenne, 2), fill=row_fill, font=Font(bold=True))
        wc(r, 7, mention,       fill=row_fill, font=Font(bold=True))
        wc(r, 8, trend_txt,     fill=row_fill)
        ws.row_dimensions[r].height = 18

    # Ligne stats
    last_r = len(df_result) + 4
    ws.merge_cells(f"A{last_r}:B{last_r}")
    wc(last_r, 1, "STATISTIQUES CLASSE", fill=gold_fill, font=Font(bold=True, color="7B5800"))
    wc(last_r, 3, f"Min: {df_result['Devoir1'].min()}", fill=gold_fill)
    wc(last_r, 4, f"Min: {df_result['Devoir2'].min()}", fill=gold_fill)
    wc(last_r, 5, f"Min: {df_result['Devoir3'].min()}", fill=gold_fill)
    wc(last_r, 6, round(df_result['Moyenne'].mean(), 2), fill=gold_fill, font=Font(bold=True, color="7B5800"))

    # Largeurs
    for col, width in zip("ABCDEFGH", [6, 28, 12, 12, 12, 12, 16, 18]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

def _ar(text):
    """Convertit texte arabe pour affichage correct dans PDF (RTL + reshaping)."""
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)
    except Exception:
        return str(text)

def _clean(text):
    """Supprime les emojis et caractères non-latin pour PDF."""
    import re
    # Remplacer emojis tendance par texte
    replacements = {
        "📈": "(+)", "📉": "(-)", "➡️": "(=)",
        "🥇": "1er", "🥈": "2eme", "🥉": "3eme",
        "🎓": "", "📊": "", "📄": "",
    }
    for emoji, txt in replacements.items():
        text = text.replace(emoji, txt)
    # Supprimer emojis restants
    text = re.sub(r'[^-éàèùâêîôûäëïöüçœæ؀-ۿ -~+\-./()%:,0-9 ]', '', text)
    return text.strip()

def make_pdf_bulletin(df_result):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os, urllib.request

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    # ── Police Arabic ──
    arabic_font = "Helvetica"
    arabic_font_bold = "Helvetica-Bold"
    font_dir = "/tmp/fonts"
    os.makedirs(font_dir, exist_ok=True)
    amiri_path      = f"{font_dir}/Amiri-Regular.ttf"
    amiri_bold_path = f"{font_dir}/Amiri-Bold.ttf"

    try:
        if not os.path.exists(amiri_path):
            urllib.request.urlretrieve(
                "https://github.com/alif-type/amiri/raw/main/Amiri-Regular.ttf",
                amiri_path)
        if not os.path.exists(amiri_bold_path):
            urllib.request.urlretrieve(
                "https://github.com/alif-type/amiri/raw/main/Amiri-Bold.ttf",
                amiri_bold_path)
        pdfmetrics.registerFont(TTFont("Amiri",     amiri_path))
        pdfmetrics.registerFont(TTFont("AmiriBold", amiri_bold_path))
        arabic_font      = "Amiri"
        arabic_font_bold = "AmiriBold"
    except Exception:
        pass  # Fallback Helvetica si téléchargement échoue

    BLUE  = colors.HexColor("#1565C0")
    BLUE2 = colors.HexColor("#0D47A1")
    SKY   = colors.HexColor("#E3F2FD")
    GOLD  = colors.HexColor("#FFF8E1")
    GREEN = colors.HexColor("#E8F5E9")
    RED   = colors.HexColor("#FFEBEE")

    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("t",  fontSize=14, fontName="Helvetica-Bold", textColor=colors.white)
    sub_s   = ParagraphStyle("s",  fontSize=9,  fontName="Helvetica",      textColor=colors.HexColor("#B0BEC5"))
    stat_s  = ParagraphStyle("st", fontSize=10, fontName="Helvetica-Bold", textColor=BLUE2, spaceAfter=4)
    ar_s    = ParagraphStyle("ar", fontSize=9,  fontName=arabic_font,      alignment=2)  # RTL

    story = []

    # ── Header ──
    ht = Table([[
        Paragraph("Bulletin de Notes — Suivi de Progression", title_s),
        Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", sub_s)
    ]], colWidths=[18*cm, 9*cm])
    ht.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), BLUE),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("LEFTPADDING",(0,0),(-1,-1),16),
        ("TOPPADDING",(0,0),(-1,-1),14),
        ("BOTTOMPADDING",(0,0),(-1,-1),14),
    ]))
    story.append(ht)
    story.append(Spacer(1, 0.4*cm))

    # ── Stats ──
    moy_classe = df_result["Moyenne"].mean()
    nb_reussi  = (df_result["Moyenne"] >= 10).sum()
    nb_echec   = (df_result["Moyenne"] < 10).sum()
    story.append(Paragraph(
        f"<b>Classe :</b> {len(df_result)} etudiants  |  "
        f"<b>Moyenne classe :</b> {moy_classe:.2f}/20  |  "
        f"<b>Reussite :</b> {nb_reussi} ({nb_reussi/len(df_result)*100:.0f}%)  |  "
        f"<b>Echec :</b> {nb_echec}", stat_s))
    story.append(Spacer(1, 0.3*cm))

    # ── Détection Activités ──
    has_act = "Activites" in df_result.columns and df_result["Activites"].notna().any()

    # ── En-têtes tableau ──
    if has_act:
        headers = ["N°", "Nom Etudiant", "D1/20", "D2/20", "D3/20",
                   "Act./20", "Moy./20", "Mention", "Tendance", "Prog D1-D3"]
        col_w = [0.8*cm, 5.5*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 2*cm, 2.8*cm, 2.5*cm, 2*cm]
    else:
        headers = ["N°", "Nom Etudiant", "D1/20", "D2/20", "D3/20",
                   "Moy./20", "Mention", "Tendance", "Prog D1-D3"]
        col_w = [0.8*cm, 6*cm, 2*cm, 2*cm, 2*cm, 2.2*cm, 3*cm, 3*cm, 2.5*cm]

    tdata = [headers]

    df_sorted = df_result.sort_values("Moyenne", ascending=False).reset_index(drop=True)
    for i, row in enumerate(df_sorted.itertuples(), 1):
        mention, _ = get_mention(row.Moyenne)
        prog       = row.Devoir3 - row.Devoir1
        prog_str   = f"+{prog:.1f}" if prog > 0 else f"{prog:.1f}"

        # Tendance sans emoji
        diff = row.Devoir3 - row.Devoir1
        if diff > 1:    trend = "En progres"
        elif diff < -1: trend = "En baisse"
        else:           trend = "Stable"

        # Nom arabe correctement formé
        nom_ar = _ar(row.Etudiant)

        if has_act:
            act_val = f"{row.Activites:.2f}" if pd.notna(row.Activites) else "-"
            tdata.append([str(i), nom_ar,
                str(row.Devoir1), str(row.Devoir2), str(row.Devoir3),
                act_val, f"{row.Moyenne:.2f}", mention, trend, prog_str])
        else:
            tdata.append([str(i), nom_ar,
                str(row.Devoir1), str(row.Devoir2), str(row.Devoir3),
                f"{row.Moyenne:.2f}", mention, trend, prog_str])

    # Ligne moyenne
    empty = ["—"] * (len(headers) - 3)
    tdata.append(["—", "MOYENNE CLASSE", "—", "—", "—"] + empty + [f"{moy_classe:.2f}", "—", "—", "—"]
                 if not has_act else
                 ["—", "MOYENNE CLASSE", "—", "—", "—", "—", f"{moy_classe:.2f}", "—", "—", "—"])

    t = Table(tdata, colWidths=col_w, repeatRows=1)

    ts_list = [
        ("BACKGROUND",    (0,0),  (-1,0),  BLUE2),
        ("TEXTCOLOR",     (0,0),  (-1,0),  colors.white),
        ("FONTNAME",      (0,0),  (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),  (-1,-1), 8),
        ("ALIGN",         (0,0),  (-1,-1), "CENTER"),
        ("GRID",          (0,0),  (-1,-1), 0.4, colors.HexColor("#CFD8DC")),
        ("TOPPADDING",    (0,0),  (-1,-1), 5),
        ("BOTTOMPADDING", (0,0),  (-1,-1), 5),
        ("LEFTPADDING",   (0,0),  (-1,-1), 4),
        ("BACKGROUND",    (0,-1), (-1,-1), GOLD),
        ("FONTNAME",      (0,-1), (-1,-1), "Helvetica-Bold"),
        ("TEXTCOLOR",     (0,-1), (-1,-1), colors.HexColor("#7B5800")),
        # Colonne noms : police arabe + alignement droite
        ("FONTNAME",      (1,1),  (1,-2),  arabic_font),
        ("ALIGN",         (1,1),  (1,-2),  "RIGHT"),
    ]

    for i, row in enumerate(df_sorted.itertuples(), 1):
        if row.Moyenne >= 14:   fill = GREEN
        elif row.Moyenne >= 10: fill = SKY
        else:                   fill = RED
        ts_list.append(("BACKGROUND", (0,i), (-1,i), fill))

    t.setStyle(TableStyle(ts_list))
    story.append(t)
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph(
        "<font color='#546E7A' size='7'>Application Suivi des Notes v1.0 — Genere automatiquement</font>",
        styles["Normal"]))

    doc.build(story)
    buf.seek(0)
    return buf

# ── Parsers fichiers ─────────────────────────────────────────────────────────
def _detect_format(df_raw):
    """Détecte le format (MASSAR ou standard) depuis un DataFrame pandas brut."""
    # Convertir tout en string pour chercher les marqueurs arabes
    for i, row in df_raw.iterrows():
        row_str = " ".join([str(v) for v in row if pd.notna(v)])
        if "الفرض الأول" in row_str or "إسم التلميذ" in row_str:
            # Format MASSAR trouvé
            data = []
            for j in range(i + 2, len(df_raw)):
                r = df_raw.iloc[j]
                try:
                    nom = r.iloc[3] if len(r) > 3 else None
                    if nom and isinstance(nom, str) and len(nom.strip()) > 1:
                        d1  = r.iloc[6]  if len(r) > 6  else None
                        d2  = r.iloc[8]  if len(r) > 8  else None
                        d3  = r.iloc[10] if len(r) > 10 else None
                        act = r.iloc[12] if len(r) > 12 else None
                        if any(pd.notna(v) for v in [d1, d2]):
                            data.append({"Etudiant": nom.strip(),
                                         "Devoir1": d1, "Devoir2": d2,
                                         "Devoir3": d3, "Activites": act})
                except Exception:
                    continue
            if data:
                return pd.DataFrame(data)

    # Format standard : chercher ligne d'en-têtes
    for i, row in df_raw.iterrows():
        row_str = " ".join([str(v).lower() for v in row if pd.notna(v)])
        if any(x in row_str for x in ["etudiant","nom","eleve","student","name"]):
            df = df_raw.iloc[i+1:].copy()
            df.columns = [str(df_raw.iloc[i, c]).strip() if pd.notna(df_raw.iloc[i, c])
                          else f"col_{c}" for c in range(len(df.columns))]
            col_map = {}
            for c in df.columns:
                cl = str(c).lower()
                if any(x in cl for x in ["nom","etudiant","eleve","student","name"]):
                    col_map[c] = "Etudiant"
                elif any(x in cl for x in ["d1","devoir1","devoir 1","note1","ds1","fard1"]):
                    col_map[c] = "Devoir1"
                elif any(x in cl for x in ["d2","devoir2","devoir 2","note2","ds2","fard2"]):
                    col_map[c] = "Devoir2"
                elif any(x in cl for x in ["d3","devoir3","devoir 3","note3","ds3","fard3"]):
                    col_map[c] = "Devoir3"
            return df.rename(columns=col_map)

    # Fallback positionnel
    if df_raw.shape[1] >= 4:
        df = df_raw.iloc[1:, :4].copy()
        df.columns = ["Etudiant","Devoir1","Devoir2","Devoir3"]
        return df
    return df_raw

def _parse_standard(df_raw):
    """Parse CSV standard."""
    for i, row in df_raw.iterrows():
        row_str = " ".join([str(v).lower() for v in row if pd.notna(v)])
        if any(x in row_str for x in ["etudiant","nom","eleve","student"]):
            df = df_raw.iloc[i+1:].copy()
            df.columns = range(len(df.columns))
            return pd.DataFrame({
                "Etudiant": df.iloc[:,0], "Devoir1": df.iloc[:,1],
                "Devoir2":  df.iloc[:,2], "Devoir3": df.iloc[:,3]
            })
    # Fallback positionnelle
    df = df_raw.copy()
    df.columns = ["Etudiant","Devoir1","Devoir2","Devoir3"] + list(range(max(0, len(df.columns)-4)))
    return df

# ── HEADER ────────────────────────────────────────────────────────────────────
# Keepalive JS : évite le timeout "Received no response from server"
st.markdown("""
<script>
// Ping le serveur toutes les 30s pour éviter le timeout
setInterval(function() {
    fetch(window.location.href, {method: 'HEAD'}).catch(()=>{});
}, 30000);
// Reconnexion automatique si perte de connexion
window.addEventListener('focus', function() {
    setTimeout(function() {
        const ws = window.parent.document.querySelector('iframe');
        if (ws) ws.src = ws.src;
    }, 500);
});
</script>
<div class="app-header">
    <div class="app-tag">Outil Pédagogique</div>
    <div class="app-title">🎓 Suivi des Notes Étudiants</div>
    <div class="app-sub">Analyse de progression — Import Excel / CSV</div>
</div>
""", unsafe_allow_html=True)

# ── SESSION STATE ──────────────────────────────────────────────────────────────
if "df_loaded" not in st.session_state:
    st.session_state.df_loaded   = None
if "file_id" not in st.session_state:
    st.session_state.file_id     = None

# ── IMPORT FICHIER ─────────────────────────────────────────────────────────────
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">📂 Import du fichier de notes</div>', unsafe_allow_html=True)
uploaded = st.file_uploader("Choisir un fichier Excel ou CSV",
                            type=["xlsx", "xls", "csv"])
st.markdown('</div>', unsafe_allow_html=True)

# ── TRAITEMENT ─────────────────────────────────────────────────────────────────
if uploaded is not None:
    # Identifiant unique du fichier : nom + taille
    file_id = f"{uploaded.name}_{uploaded.size}"

    # Ne retraiter que si c'est un nouveau fichier
    if st.session_state.file_id != file_id:
        with st.spinner("Lecture du fichier..."):
            try:
                import warnings
                file_bytes = uploaded.getvalue()
                fname = uploaded.name.lower()

                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    if fname.endswith(".csv"):
                        try:
                            df_raw = pd.read_csv(io.BytesIO(file_bytes), header=None, encoding="utf-8")
                        except Exception:
                            df_raw = pd.read_csv(io.BytesIO(file_bytes), header=None, encoding="latin-1")
                        df_raw.columns = range(len(df_raw.columns))
                        df_new = _parse_standard(df_raw)
                    elif fname.endswith(".xls"):
                        df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, engine="xlrd")
                        df_new = _detect_format(df_raw)
                    else:
                        xl = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
                        sheet = "NotesCC" if "NotesCC" in xl.sheet_names else xl.sheet_names[0]
                        df_raw = xl.parse(sheet, header=None)
                        df_new = _detect_format(df_raw)

                # Sauvegarder en session
                st.session_state.df_loaded = df_new
                st.session_state.file_id   = file_id

            except Exception as e:
                st.error(f"Erreur de lecture : {type(e).__name__} — {e}")
                st.session_state.df_loaded = None
                st.session_state.file_id   = None



if uploaded is not None and st.session_state.df_loaded is not None:
    df = st.session_state.df_loaded
    required = ["Etudiant","Devoir1","Devoir2"]
    missing  = [c for c in required if c not in df.columns]
    if missing:
        st.error("Colonnes introuvables. Vérifiez le format du fichier.")
    else:
            # Ajouter colonnes manquantes
            if "Devoir3"   not in df.columns: df["Devoir3"]   = None
            if "Activites" not in df.columns: df["Activites"] = None

            cols_base = ["Etudiant","Devoir1","Devoir2","Devoir3","Activites"]
            df = df[[c for c in cols_base if c in df.columns]].copy()
            df = df.dropna(subset=["Etudiant"])
            df["Devoir1"]   = pd.to_numeric(df["Devoir1"],   errors="coerce")
            df["Devoir2"]   = pd.to_numeric(df["Devoir2"],   errors="coerce")
            df["Devoir3"]   = pd.to_numeric(df["Devoir3"],   errors="coerce")
            df["Activites"] = pd.to_numeric(df["Activites"], errors="coerce")
            df = df.dropna(subset=["Devoir1","Devoir2"])
            df["Etudiant"] = df["Etudiant"].astype(str).str.strip()
            df = df[df["Etudiant"].str.len() > 1]

            # Détecter le nombre de devoirs réels
            has_d3  = df["Devoir3"].notna().any()
            has_act = df["Activites"].notna().any()
            nb_dev  = 3 if has_d3 else 2  # 2 ou 3 devoirs

            # Moyenne selon nb_dev et présence activités
            def calc_moyenne(r):
                notes = [r["Devoir1"], r["Devoir2"]]
                if has_d3 and pd.notna(r["Devoir3"]):
                    notes.append(r["Devoir3"])
                if has_act and pd.notna(r["Activites"]):
                    notes.append(r["Activites"])
                return sum(notes) / len(notes)
            df["Moyenne"] = df.apply(calc_moyenne, axis=1)

            # Progression : D2-D1 si 2 devoirs, D3-D1 si 3 devoirs
            if has_d3:
                df["Progression"] = df["Devoir3"] - df["Devoir1"]
            else:
                df["Progression"] = df["Devoir2"] - df["Devoir1"]

            # ── STATS GLOBALES ──────────────────────────────────────────────
            moy_classe  = df["Moyenne"].mean()
            nb_reussi   = (df["Moyenne"] >= 10).sum()
            nb_echec    = (df["Moyenne"] < 10).sum()
            meilleur    = df.loc[df["Moyenne"].idxmax(), "Etudiant"]
            plus_progres= df.loc[df["Progression"].idxmax(), "Etudiant"]
            moy_d1  = df["Devoir1"].mean()
            moy_d2  = df["Devoir2"].mean()
            moy_d3  = df["Devoir3"].mean()  if has_d3  else None
            moy_act = df["Activites"].mean() if has_act else None

            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<div class="card-title">Statistiques Generales</div>', unsafe_allow_html=True)

            nb_total   = len(df)
            pct_reussi = int(nb_reussi / nb_total * 100)
            pct_echec  = int(nb_echec  / nb_total * 100)

            html_row1 = (
                '<div class="stat-grid">'
                + '<div class="stat-box"><div class="stat-label">Etudiants</div>'
                + '<div class="stat-value">' + str(nb_total) + '</div></div>'
                + '<div class="stat-box gold"><div class="stat-label">Moyenne classe</div>'
                + '<div class="stat-value">' + f"{moy_classe:.2f}" + '/20</div></div>'
                + '<div class="stat-box green"><div class="stat-label">Reussite &ge;10</div>'
                + '<div class="stat-value">' + str(nb_reussi) + ' (' + str(pct_reussi) + '%)</div></div>'
                + '<div class="stat-box red"><div class="stat-label">Echec &lt;10</div>'
                + '<div class="stat-value">' + str(nb_echec) + ' (' + str(pct_echec) + '%)</div></div>'
                + '</div>'
            )
            st.markdown(html_row1, unsafe_allow_html=True)

            act_box = (
                '<div class="stat-box gold"><div class="stat-label">Moy. Activites</div>'
                + '<div class="stat-value">' + f"{moy_act:.2f}" + '</div></div>'
            ) if moy_act is not None else ""

            html_row2 = (
                '<div class="stat-grid">'
                + '<div class="stat-box"><div class="stat-label">Moy. Devoir 1</div>'
                + '<div class="stat-value">' + f"{moy_d1:.2f}" + '</div></div>'
                + '<div class="stat-box"><div class="stat-label">Moy. Devoir 2</div>'
                + '<div class="stat-value">' + f"{moy_d2:.2f}" + '</div></div>'
                + ('<div class="stat-box"><div class="stat-label">Moy. Devoir 3</div>'
                + '<div class="stat-value">' + f"{moy_d3:.2f}" + '</div></div>') if moy_d3 is not None else ''
                + act_box
                + '</div>'
            )
            st.markdown(html_row2, unsafe_allow_html=True)

            c1, c2 = st.columns(2)
            c1.metric("Meilleur etudiant", meilleur, f"{df['Moyenne'].max():.2f}/20")
            c2.metric("Plus grande progression", plus_progres,
                      f"+{df['Progression'].max():.2f} pts")
            st.markdown('</div>', unsafe_allow_html=True)

            # ── TABS ───────────────────────────────────────────────────────
            tab1, tab2, tab3, tab4 = st.tabs([
                "📋 Tableau de bord", "📈 Graphiques", "🏆 Classement", "📄 Export"
            ])

            # ── TAB 1 : Tableau de bord ─────────────────────────────────────
            with tab1:
                st.markdown('<div class="card">', unsafe_allow_html=True)
                st.markdown('<div class="card-title">📋 Détail par étudiant</div>', unsafe_allow_html=True)

                for _, row in df.sort_values("Moyenne", ascending=False).iterrows():
                    mention, mention_cls = get_mention(row["Moyenne"])
                    trend_txt, trend_color = get_trend(row["Devoir1"], row["Devoir2"], row["Devoir3"])
                    prog     = row["Devoir3"] - row["Devoir1"]
                    prog_str = ("+{:.1f}".format(prog)) if prog > 0 else "{:.1f}".format(prog)
                    prog_color = "#2E7D32" if prog > 0 else "#C62828"
                    moy_str  = "{:.2f}".format(row["Moyenne"])
                    nom      = str(row["Etudiant"])
                    d1       = str(row["Devoir1"])
                    d2       = str(row["Devoir2"])
                    d3       = str(row["Devoir3"]) if has_d3 and pd.notna(row["Devoir3"]) else None
                    pbar     = progress_bar(row["Moyenne"], 20, trend_color)

                    # Activités — gérer note manquante
                    act_val = row["Activites"] if "Activites" in row.index else None
                    if act_val is not None and pd.notna(act_val):
                        act_html = (
                            '&nbsp;&nbsp; Act: <b style="color:#7B5800">'
                            + str(act_val) + '</b>'
                        )
                    else:
                        act_html = ""

                    d3_part  = (' &nbsp;&rarr;&nbsp; D3: <b>' + str(d3) + '</b>') if (has_d3 and d3 and str(d3) not in ['None','nan']) else ''
                    prog_lbl = '&nbsp;&nbsp; Prog D1-D3 : ' if has_d3 else '&nbsp;&nbsp; Prog D1-D2 : '
                    html = (
                        '<div style="background:#F8FAFC;border:1px solid #E0E7EF;'
                        'border-radius:12px;padding:14px 18px;margin-bottom:10px;'
                        'display:flex;justify-content:space-between;'
                        'align-items:center;flex-wrap:wrap;gap:10px">'
                        '<div style="flex:1">'
                        '<div style="font-weight:800;font-size:15px;color:#0D1B2A">' + nom + '</div>'
                        '<div style="font-size:12px;color:#546E7A;margin-top:3px">'
                        'D1: <b>' + d1 + '</b> &nbsp;&rarr;&nbsp; D2: <b>' + d2 + '</b>'
                        + d3_part + act_html + prog_lbl
                        + '<b style="color:' + prog_color + '">' + prog_str + ' pts</b>'
                        '</div>'
                        + pbar
                        + '</div>'
                        '<div style="text-align:right">'
                        '<div style="font-size:24px;font-weight:900;color:#0D47A1">'
                        + moy_str + '<span style="font-size:12px;color:#546E7A">/20</span></div>'
                        '<span class="mention ' + mention_cls + '">' + mention + '</span>'
                        '<div style="font-size:12px;margin-top:4px">' + trend_txt + '</div>'
                        '</div></div>'
                    )
                    st.markdown(html, unsafe_allow_html=True)

                st.markdown('</div>', unsafe_allow_html=True)

            # ── TAB 2 : Graphiques ──────────────────────────────────────────
            with tab2:
                st.markdown('<div class="card">', unsafe_allow_html=True)
                st.markdown('<div class="card-title">🏅 Répartition des mentions</div>', unsafe_allow_html=True)

                # Camembert mentions avec matplotlib
                import matplotlib.pyplot as plt
                import matplotlib.patches as mpatches

                mentions_count = {}
                for _, row in df.iterrows():
                    m, _ = get_mention(row["Moyenne"])
                    mentions_count[m] = mentions_count.get(m, 0) + 1

                ordre    = ["Excellent","Bien","Assez Bien","Passable","Insuffisant"]
                couleurs = {"Excellent":"#2E7D32","Bien":"#1565C0",
                            "Assez Bien":"#F9A825","Passable":"#E65100","Insuffisant":"#C62828"}
                labels   = [m for m in ordre if m in mentions_count]
                sizes    = [mentions_count[m] for m in labels]
                colors   = [couleurs[m] for m in labels]

                fig, ax = plt.subplots(figsize=(6, 5))
                fig.patch.set_facecolor("#FFFFFF")
                ax.set_facecolor("#FFFFFF")
                wedges, texts, autotexts = ax.pie(
                    sizes, labels=None, colors=colors,
                    autopct=lambda p: f"{p:.1f}%\n({int(round(p*sum(sizes)/100))})",
                    startangle=90, pctdistance=0.75,
                    wedgeprops=dict(width=0.55, edgecolor="white", linewidth=2))
                for at in autotexts:
                    at.set_fontsize(11)
                    at.set_fontweight("bold")
                    at.set_color("white")
                # Légende
                legend_patches = [mpatches.Patch(color=couleurs[m], label=f"{m}  ({mentions_count[m]})")
                                  for m in labels]
                ax.legend(handles=legend_patches, loc="lower center",
                          bbox_to_anchor=(0.5, -0.12), ncol=3,
                          fontsize=10, frameon=False)
                ax.set_title("Répartition des mentions", fontsize=14,
                             fontweight="bold", color="#0D47A1", pad=16)
                plt.tight_layout()
                col_pie, _ = st.columns([2, 1])
                with col_pie:
                    st.pyplot(fig, use_container_width=True)
                plt.close(fig)
                st.markdown('</div>', unsafe_allow_html=True)

            # ── TAB 3 : Classement ──────────────────────────────────────────
            with tab3:
                st.markdown('<div class="card">', unsafe_allow_html=True)
                st.markdown('<div class="card-title">🏆 Classement général</div>', unsafe_allow_html=True)

                df_rank = df.sort_values("Moyenne", ascending=False).reset_index(drop=True)
                df_rank["Rang"]      = df_rank.index + 1
                df_rank["Mention"]   = df_rank["Moyenne"].apply(lambda x: get_mention(x)[0])
                # Progression = D3 - D1 uniquement (3 premiers devoirs)
                if has_d3:
                    prog_series = df_rank["Devoir3"] - df_rank["Devoir1"]
                else:
                    prog_series = df_rank["Devoir2"] - df_rank["Devoir1"]
                df_rank["Prog D1-D2/D3"] = prog_series.apply(
                    lambda x: ("+{:.1f}".format(x)) if x > 0 else "{:.1f}".format(x))
                if has_d3:
                    df_rank["Tendance"] = df_rank.apply(
                        lambda r: get_trend(r["Devoir1"], r["Devoir2"], r["Devoir3"])[0], axis=1)
                else:
                    df_rank["Tendance"] = df_rank.apply(
                        lambda r: get_trend(r["Devoir1"], r["Devoir2"], r["Devoir2"])[0], axis=1)
                df_rank["Moy."]      = df_rank["Moyenne"].round(2)
                has_act = "Activites" in df_rank.columns and df_rank["Activites"].notna().any()
                has_d3_rank = "Devoir3" in df_rank.columns and df_rank["Devoir3"].notna().any()
                has_act_rank = has_act

                # ── Tableau HTML complet sans scrolling ──
                act_header = "<th>Activités</th>" if has_act else ""
                rows_html  = ""
                for _, r in df_rank.iterrows():
                    rang = int(r["Rang"])
                    if rang == 1:   medal = "🥇"
                    elif rang == 2: medal = "🥈"
                    elif rang == 3: medal = "🥉"
                    else:           medal = str(rang)

                    moy = r["Moy."]
                    if moy >= 14:   bg = "#E8F5E9"
                    elif moy >= 10: bg = "#E3F2FD"
                    else:           bg = "#FFEBEE"

                    prog_val = r["Prog D1→D3"]
                    prog_color = "#2E7D32" if "+" in str(prog_val) else "#C62828"
                    mention, m_cls = get_mention(moy)

                    if has_act:
                        act_v = r["Activites"]
                        act_display = str(act_v) if pd.notna(act_v) else "-"
                        act_cell = '<td style="text-align:center;font-weight:700;color:#7B5800">' + act_display + '</td>'
                    else:
                        act_cell = ""
                    _ = act_cell  # éviter warning

                    rows_html += (
                        '<tr style="background:' + bg + '">'
                        '<td style="text-align:center;font-weight:800;font-size:16px">' + str(medal) + '</td>'
                        '<td style="font-weight:700;padding-left:8px">' + str(r["Etudiant"]) + '</td>'
                        '<td style="text-align:center">' + str(r["Devoir1"]) + '</td>'
                        '<td style="text-align:center">' + str(r["Devoir2"]) + '</td>'
                        + ('<td style="text-align:center">' + (str(r["Devoir3"]) if pd.notna(r["Devoir3"]) else '-') + '</td>' if has_d3 else '')
                        + act_cell +
                        '<td style="text-align:center;font-weight:900;color:#0D47A1;font-size:16px">' + str(moy) + '</td>'
                        '<td style="text-align:center"><span style="background:white;border-radius:12px;padding:2px 10px;font-size:11px;font-weight:700">' + mention + '</span></td>'
                        '<td style="text-align:center;font-weight:700;color:' + prog_color + '">' + str(prog_val) + '</td>'
                        '<td style="text-align:center">' + str(r["Tendance"]) + '</td>'
                        '</tr>'
                    )

                table_html = (
                    '<table style="width:100%;border-collapse:collapse;font-size:13px">'
                    '<thead><tr style="background:#0D47A1;color:white">'
                    '<th style="padding:10px 6px">Rang</th>'
                    '<th style="padding:10px 8px;text-align:left">Etudiant</th>'
                    '<th style="padding:10px 6px">D1</th>'
                    '<th style="padding:10px 6px">D2</th>'
                    + ('<th style="padding:10px 6px">D3</th>' if has_d3 else '')
                    + act_header +
                    '<th style="padding:10px 6px">Moy.</th>'
                    '<th style="padding:10px 6px">Mention</th>'
                    + '<th style="padding:10px 6px">' + prog_lbl + '</th>'
                    '<th style="padding:10px 6px">Tendance</th>'
                    '</tr></thead>'
                    '<tbody>' + rows_html + '</tbody>'
                    '</table>'
                )
                st.markdown(table_html, unsafe_allow_html=True)

                # Top 3
                st.markdown("---")
                st.markdown("**🥇 Top 3 étudiants**")
                top3_cols = st.columns(3)
                medals_top = ["🥇","🥈","🥉"]
                for idx, (medal, col) in enumerate(zip(medals_top, top3_cols)):
                    if idx < len(df_rank):
                        r = df_rank.iloc[idx]
                        mention, _ = get_mention(float(r["Moy."]))
                        col.metric(f"{medal} {r['Etudiant']}", f"{r['Moy.']}/20", mention)

                st.markdown('</div>', unsafe_allow_html=True)

            # ── TAB 4 : Export ──────────────────────────────────────────────
            with tab4:
                st.markdown('<div class="card">', unsafe_allow_html=True)
                st.markdown('<div class="card-title">📄 Exporter le bulletin complet</div>', unsafe_allow_html=True)

                st.info("Les exports incluent : toutes les notes, moyennes, mentions, classement et progression.")

                dl1, dl2 = st.columns(2)
                with dl1:
                    st.markdown('<div class="dl-excel">', unsafe_allow_html=True)
                    try:
                        xlsx = make_excel_bulletin(df)
                        st.download_button("📊 Télécharger Excel",
                            data=xlsx,
                            file_name=f"Bulletin_Notes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_excel_notes")
                    except ImportError:
                        st.info("Ajoutez `openpyxl` à requirements.txt")
                    st.markdown('</div>', unsafe_allow_html=True)

                with dl2:
                    st.markdown('<div class="dl-pdf">', unsafe_allow_html=True)
                    try:
                        pdf = make_pdf_bulletin(df)
                        st.download_button("📄 Télécharger PDF",
                            data=pdf,
                            file_name=f"Bulletin_Notes_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                            mime="application/pdf",
                            key="dl_pdf_notes")
                    except ImportError:
                        st.info("Ajoutez `reportlab` à requirements.txt")
                    st.markdown('</div>', unsafe_allow_html=True)

                st.markdown('</div>', unsafe_allow_html=True)



else:
    st.markdown("""
    <div style="text-align:center;padding:48px 20px;color:#546E7A">
        <div style="font-size:56px;margin-bottom:16px">📂</div>
        <div style="font-size:18px;font-weight:700;color:#0D47A1;margin-bottom:8px">
            Importez votre fichier de notes
        </div>
        <div style="font-size:14px">
            Formats acceptes : <b>Excel (.xlsx, .xls)</b> ou <b>CSV (.csv)</b>
        </div>
    </div>
    """, unsafe_allow_html=True)

# ── Footer ─────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="app-footer">
    <span class="badge">Suivi Notes v1.0</span> &nbsp;|&nbsp;
    Analyse de progression étudiante &nbsp;|&nbsp;
    {datetime.now().strftime('%Y')}
</div>
""", unsafe_allow_html=True)
